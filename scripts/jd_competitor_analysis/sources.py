"""按表头确认并读取京东竞品分析工作簿。"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .input_files import WorkbookInput, extract_dates


LOGGER = logging.getLogger(__name__)

ROLE_RULES = {
    "self_real": {
        "label": "本品真实 SPU 数据",
        "hints": ("经营状况-商品明细", "商品明细", "spu真实数据"),
        "required": ("时间", "SPU", "访客数", "成交金额"),
        "level": "core",
    },
    "core": {
        "label": "竞品核心数据对比",
        "hints": ("核心指标", "竞品数据对比"),
        "required": ("时间", "本品访客数"),
        "level": "core",
    },
    "traffic": {
        "label": "流量来源对比",
        "hints": ("流量来源", "流量来源对比"),
        "required": ("时间", "一级渠道", "本品访客数"),
        "level": "core",
    },
    "keywords": {
        "label": "引流关键词对比",
        "hints": ("关键词榜单", "引流关键词对比"),
        "required": ("日期", "关键词", "SPUID", "访客数", "成交金额"),
        "level": "complete",
    },
    "customer_profile": {
        "label": "成交客户画像对比",
        "hints": ("成交客户画像", "成交客户对比"),
        "required": ("时间", "画像类型", "本品成交客户数占比"),
        "level": "complete",
    },
    "promotion": {
        "label": "推广数据对比",
        "hints": ("广告流量对比", "推广数据对比"),
        "required": (),
        "level": "enhancement",
    },
}


def clean_text(value: Any) -> str:
    """把单元格值转换为去除首尾空白的文本。"""

    return "" if value is None else str(value).strip()


def clean_identifier(value: Any) -> str:
    """把 Excel 中的数字标识统一为无小数文本。"""

    text = clean_text(value)
    return text[:-2] if text.endswith(".0") else text


def workbook_headers(path: Path) -> list[tuple[str, list[str]]]:
    """读取工作簿每个工作表的首行表头。"""

    workbook = load_workbook(path, read_only=False, data_only=True)
    result = []
    for sheet in workbook.worksheets:
        headers = [clean_text(sheet.cell(1, column).value) for column in range(1, sheet.max_column + 1)]
        result.append((sheet.title, headers))
    workbook.close()
    return result


def _header_requirements(
    role: str,
    rule: dict[str, Any],
    competitor_prefix: str,
) -> tuple[tuple[str, ...], tuple[tuple[str, ...], ...]]:
    """生成一个数据角色的完整表头条件。"""

    required = tuple(rule["required"])
    any_groups: tuple[tuple[str, ...], ...] = ()
    if role in {"core", "traffic"}:
        required += (f"{competitor_prefix}访客数",)
    elif role == "customer_profile":
        required += (f"{competitor_prefix}成交客户数占比",)
    elif role == "promotion":
        any_groups = (
            (
                "非全站-本店商品广告点击数",
                "非全站-本店商品广告总订单金额",
                "全站-本店商品核心位置点击数",
                "全站-本店商品全站交易额",
            ),
            (
                f"非全站-{competitor_prefix}广告点击数",
                f"非全站-{competitor_prefix}广告总订单金额",
                f"全站-{competitor_prefix}核心位置点击数",
                f"全站-{competitor_prefix}全站交易额",
            ),
        )
    return required, any_groups


def _headers_match(
    headers: list[str],
    required: tuple[str, ...],
    any_groups: tuple[tuple[str, ...], ...],
) -> bool:
    """判断工作表表头是否满足一个数据角色。"""

    header_set = set(headers)
    return all(header in header_set for header in required) and all(
        any(header in header_set for header in group) for group in any_groups
    )


def _matching_sheets(
    workbooks: list[WorkbookInput],
    headers_by_path: dict[Path, list[tuple[str, list[str]]]],
    required: tuple[str, ...],
    any_groups: tuple[tuple[str, ...], ...],
) -> list[tuple[WorkbookInput, str]]:
    """从候选工作簿中找出满足表头条件的工作表。"""

    matches: list[tuple[WorkbookInput, str]] = []
    for workbook in workbooks:
        for sheet_name, headers in headers_by_path[workbook.path]:
            if _headers_match(headers, required, any_groups):
                matches.append((workbook, sheet_name))
    return matches


def discover_sources(workbooks: list[WorkbookInput], competitor_prefix: str) -> dict[str, dict[str, Any]]:
    """发现并确认六类真实源文件。

    功能说明：先按原始文件名语义缩小候选集，再以工作簿表头唯一确认数据角色和工作表。
    参数 workbooks：当前周期直接 XLSX 与 ZIP 内 XLSX 的统一输入列表。
    参数 competitor_prefix：目标竞品字段前缀。
    返回值：按数据角色组织的文件、工作表和完整性信息。
    """

    headers_by_path: dict[Path, list[tuple[str, list[str]]]] = {}
    for workbook in workbooks:
        LOGGER.info("读取工作簿表头：%s", workbook.file_name)
        headers_by_path[workbook.path] = workbook_headers(workbook.path)

    sources: dict[str, dict[str, Any]] = {}
    for role, rule in ROLE_RULES.items():
        required, any_groups = _header_requirements(role, rule, competitor_prefix)
        hinted = [
            workbook
            for workbook in workbooks
            if any(hint.lower() in workbook.file_name.lower() for hint in rule["hints"])
        ]
        candidates = _matching_sheets(hinted or workbooks, headers_by_path, required, any_groups)
        if hinted and not candidates:
            candidates = _matching_sheets(workbooks, headers_by_path, required, any_groups)

        if len(candidates) != 1:
            status = "missing" if not candidates else "conflict"
            candidate_names = [f"{item.file_name}/{sheet}" for item, sheet in candidates]
            message = f"{rule['label']}无法唯一匹配，候选={candidate_names}"
            if rule["level"] == "core":
                raise RuntimeError(message)
            LOGGER.warning("%s，状态=%s", message, status)
            sources[role] = {
                "role": role,
                "label": rule["label"],
                "required_level": rule["level"],
                "status": status,
                "warnings": [message],
            }
            continue

        workbook, sheet_name = candidates[0]
        sources[role] = {
            "role": role,
            "label": rule["label"],
            "required_level": rule["level"],
            "status": "ready",
            "path": workbook.path,
            "file_name": workbook.file_name,
            "sheet_name": sheet_name,
            "warnings": [],
        }
        LOGGER.info("已定位%s：%s / %s", rule["label"], workbook.file_name, sheet_name)
    return sources


def read_rows(source: dict[str, Any]) -> list[dict[str, Any]]:
    """读取已确认工作表的所有数据行。"""

    if source.get("status") != "ready":
        return []
    workbook = load_workbook(source["path"], read_only=False, data_only=True)
    sheet = workbook[source["sheet_name"]]
    headers = [clean_text(sheet.cell(1, column).value) for column in range(1, sheet.max_column + 1)]
    rows = []
    for row_index in range(2, sheet.max_row + 1):
        values = [sheet.cell(row_index, column).value for column in range(1, sheet.max_column + 1)]
        row = dict(zip(headers, values))
        row["_row_index"] = row_index
        rows.append(row)
    workbook.close()
    LOGGER.info("已读取%s：%s 行", source["label"], len(rows))
    return rows


def validate_rows_period(
    role: str,
    rows: list[dict[str, Any]],
    granularity: str,
    period_start: str,
    period_end: str,
) -> None:
    """校验工作表业务行的周期。

    功能说明：检查带时间字段的数据行是否匹配当前周期；月维度兼容京东 `YYYYMM` 格式。
    参数 role：当前数据角色。
    参数 rows：当前工作表读取出的全部业务行。
    参数 granularity：当前分析粒度。
    参数 period_start：当前周期开始日期。
    参数 period_end：当前周期结束日期。
    返回值：无；表内时间冲突时抛出 RuntimeError。
    """

    if not rows or role == "promotion":
        return
    field_name = "日期" if role == "keywords" else "时间"
    values = {clean_text(row.get(field_name)) for row in rows}
    values.discard("")
    for value in values:
        dates = extract_dates(value)
        iso_range_matched = bool(
            dates and dates[0] == period_start and dates[-1] == period_end
        )
        compact_month_matched = bool(
            granularity == "month"
            and len(value) == 6
            and value.isdigit()
            and period_start[:7] == period_end[:7]
            and value == period_start[:7].replace("-", "")
        )
        if not iso_range_matched and not compact_month_matched:
            raise RuntimeError(
                f"{ROLE_RULES[role]['label']}表内周期与目录冲突：{value}，"
                f"目录周期={period_start}_{period_end}"
            )
