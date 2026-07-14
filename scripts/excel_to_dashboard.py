from __future__ import annotations

import argparse
import json
import logging
import re
from dataclasses import asdict, dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


logger = logging.getLogger(__name__)


@dataclass(frozen=True)
class MetricDefinition:
    """定义一个核心指标的标准字段。

    功能说明：统一核心指标 ID、中文字段名和网页展示单位。
    参数 id：标准指标 ID。
    参数 label：京东导出表中的指标名称。
    参数 unit：网页使用的展示单位。
    返回值：不可变的指标定义对象。
    """

    id: str
    label: str
    unit: str = ""


@dataclass(frozen=True)
class ParsedRange:
    """保存一个已解析的区间值。

    功能说明：保留区间原文、上下界、中位数、百分比和缺失状态。
    参数 raw：原始单元格文本。
    参数 low：区间下界。
    参数 high：区间上界。
    参数 mid：区间中位数。
    参数 is_percent：是否为百分比。
    参数 missing：是否缺失或解析失败。
    返回值：可通过 `asdict` 写入 JSON 的区间对象。
    """

    raw: str | None
    low: float | None
    high: float | None
    mid: float | None
    is_percent: bool
    missing: bool


CORE_METRICS = (
    MetricDefinition("gmv", "成交金额"),
    MetricDefinition("sold_units", "成交商品件数"),
    MetricDefinition("orders", "成交单量"),
    MetricDefinition("views", "浏览量"),
    MetricDefinition("visitors", "访客数"),
    MetricDefinition("cart_users", "加购人数"),
    MetricDefinition("conversion_rate", "成交转化率", "%"),
    MetricDefinition("customer_price", "成交客单价"),
)

CORE_CARD_IDS = {"gmv", "visitors", "conversion_rate", "customer_price"}
DIMENSION_NAMES = {"性别", "年龄", "地区", "省份", "城市"}

ROLE_RULES = {
    "self_real": {
        "label": "本品真实 SPU 数据",
        "hints": ("spu真实数据", "商品明细"),
        "required": ("SPU", "访客数", "成交金额"),
        "level": "core",
    },
    "core": {
        "label": "竞品核心数据对比",
        "hints": ("竞品数据对比",),
        "required": ("本品访客数",),
        "level": "core",
    },
    "traffic": {
        "label": "流量来源对比",
        "hints": ("流量来源对比",),
        "required": ("一级渠道", "本品访客数"),
        "level": "core",
    },
    "keywords": {
        "label": "引流关键词对比",
        "hints": ("引流关键词对比",),
        "required": ("关键词", "SPUID", "访客数", "成交金额"),
        "level": "complete",
    },
    "customer_profile": {
        "label": "成交客户画像对比",
        "hints": ("成交客户对比",),
        "required": ("画像类型", "本品成交客户数占比"),
        "level": "complete",
    },
    "promotion": {
        "label": "推广数据对比",
        "hints": ("推广数据对比",),
        "required": (),
        "level": "enhancement",
    },
}

GRANULARITY_DIRS = {
    "day": "天",
    "week": "周",
    "month": "月",
}


def clean_text(value: Any) -> str:
    """把单元格值转换为去除首尾空白的文本。"""

    return "" if value is None else str(value).strip()


def clean_identifier(value: Any) -> str:
    """把 SPU 等标识符转换为不带小数后缀的文本。"""

    text = clean_text(value)
    return text[:-2] if text.endswith(".0") else text


def to_number(value: Any) -> float | None:
    """把普通数字转换为浮点数，无法转换时返回空值。"""

    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def parse_range(value: Any) -> ParsedRange:
    """解析京东导出表中的金额、数量和百分比区间。

    功能说明：支持 `万`、货币符号、百分号、区间、单值和缺失值。
    参数 value：Excel 单元格原始值。
    返回值：标准区间对象；解析失败时返回缺失结构并记录日志。
    """

    raw = clean_text(value)
    if not raw or raw == "-":
        return ParsedRange(raw=raw or None, low=None, high=None, mid=None, is_percent=False, missing=True)
    is_percent = "%" in raw
    text = raw.replace(",", "").replace("￥", "").replace("¥", "").replace(" ", "").replace("%", "")
    parts = re.split(r"~|－|—", text)
    numbers: list[float] = []
    for part in parts:
        if not part:
            continue
        multiplier = 10000.0 if "万" in part else 1.0
        try:
            numbers.append(float(part.replace("万", "")) * multiplier)
        except ValueError:
            logger.warning("区间解析失败：%s", raw)
            return ParsedRange(raw=raw, low=None, high=None, mid=None, is_percent=is_percent, missing=True)
    if not numbers:
        return ParsedRange(raw=raw, low=None, high=None, mid=None, is_percent=is_percent, missing=True)
    low = numbers[0]
    high = numbers[1] if len(numbers) > 1 else numbers[0]
    if is_percent:
        low /= 100
        high /= 100
    return ParsedRange(raw=raw, low=low, high=high, mid=(low + high) / 2, is_percent=is_percent, missing=False)


def clamp(value: float, interval: ParsedRange) -> float:
    """把数值限制在区间边界内。"""

    if interval.low is not None:
        value = max(value, interval.low)
    if interval.high is not None:
        value = min(value, interval.high)
    return value


def in_range(value: float | None, interval: ParsedRange) -> bool | None:
    """判断数值是否落入有效区间。"""

    if value is None or interval.low is None or interval.high is None:
        return None
    return interval.low <= value <= interval.high


def extract_dates(text: str) -> list[str]:
    """从文件名或周期文本中提取规范日期。"""

    dates = []
    for year, month, day in re.findall(r"(20\d{2})[-.](\d{1,2})[-.](\d{1,2})", text):
        normalized = f"{year}-{int(month):02d}-{int(day):02d}"
        if normalized not in dates:
            dates.append(normalized)
    return dates


def period_matches(path: Path, period_file: str) -> bool:
    """判断文件名日期是否与目标周期一致。"""

    target = extract_dates(period_file)
    found = extract_dates(path.stem)
    return not target or not found or target == found


def period_fields(granularity: str, period_text: str) -> dict[str, str]:
    """生成标准周期字段。

    功能说明：从周期文本中提取起止日期，并生成展示周期和唯一周期键。
    参数 granularity：分析粒度，取值为 `day`、`week` 或 `month`。
    参数 period_text：包含一个或两个日期的周期文本。
    返回值：包含 `period`、`period_start`、`period_end` 和 `period_key` 的字典。
    """

    dates = extract_dates(period_text)
    if not dates:
        raise ValueError(f"周期未包含有效日期：{period_text}")
    period_start = dates[0]
    period_end = dates[-1]
    period = period_start if granularity == "day" and period_start == period_end else f"{period_start}~{period_end}"
    return {
        "period": period,
        "period_start": period_start,
        "period_end": period_end,
        "period_key": f"{granularity}:{period_start}_{period_end}",
    }


def discover_periods(input_dir: Path) -> list[tuple[str, str]]:
    """发现粒度目录中的全部分析周期。

    功能说明：以竞品核心数据对比文件为周期锚点，提取并排序每个独立起止日期。
    参数 input_dir：某一粒度的原始 Excel 目录。
    返回值：按开始日期和结束日期升序排列的周期元组列表。
    """

    periods: set[tuple[str, str]] = set()
    for path in input_dir.glob("*.xlsx"):
        if path.name.startswith("~$") or "竞品数据对比" not in path.stem:
            continue
        dates = extract_dates(path.stem)
        if not dates:
            logger.warning("核心数据文件未识别到周期：%s", path.name)
            continue
        periods.add((dates[0], dates[-1]))
    result = sorted(periods)
    logger.info("已发现周期：%s，共 %s 个", input_dir, len(result))
    return result


def period_in_window(period_start: str, period_end: str, start_date: str | None, end_date: str | None) -> bool:
    """判断周期是否落入调用方指定的分析窗口。"""

    if start_date and period_end < start_date:
        return False
    if end_date and period_start > end_date:
        return False
    return True


def workbook_headers(path: Path) -> list[tuple[str, list[str]]]:
    """读取工作簿每个工作表的首行表头。"""

    workbook = load_workbook(path, read_only=False, data_only=True)
    result = []
    for sheet in workbook.worksheets:
        headers = [clean_text(sheet.cell(1, column).value) for column in range(1, sheet.max_column + 1)]
        result.append((sheet.title, headers))
    workbook.close()
    return result


def select_sheet(path: Path, required_headers: tuple[str, ...]) -> str:
    """按表头选择源工作表。

    功能说明：扫描所有工作表并选择满足角色字段要求的工作表。
    参数 path：Excel 文件路径。
    参数 required_headers：当前数据角色的基础必需表头。
    返回值：唯一匹配的工作表名称。
    """

    candidates = []
    for sheet_name, headers in workbook_headers(path):
        header_set = set(headers)
        if not all(header in header_set for header in required_headers):
            continue
        candidates.append(sheet_name)
    if not candidates and not required_headers:
        candidates = [item[0] for item in workbook_headers(path)]
    if len(candidates) != 1:
        raise RuntimeError(f"工作表无法唯一匹配：{path.name}，候选={candidates}")
    return candidates[0]


def discover_sources(input_dir: Path, period_file: str, competitor_prefix: str) -> dict[str, dict[str, Any]]:
    """发现并确认六类真实源文件。

    功能说明：按周期、文件角色关键词和表头定位输入文件与工作表。
    参数 input_dir：当前粒度的原始 Excel 目录。
    参数 period_file：目标周期文件名片段。
    参数 competitor_prefix：目标竞品字段前缀。
    返回值：按数据角色组织的文件、工作表和完整性信息。
    """

    files = [path for path in input_dir.glob("*.xlsx") if not path.name.startswith("~$")]
    sources: dict[str, dict[str, Any]] = {}
    for role, rule in ROLE_RULES.items():
        period_files = [path for path in files if period_matches(path, period_file)]
        hinted_files = [
            path
            for path in period_files
            if any(hint.lower() in path.stem.lower() for hint in rule["hints"])
        ]
        required_headers = tuple(rule["required"])
        if role in {"core", "traffic"}:
            required_headers += (f"{competitor_prefix}访客数",)
        scan_pool = hinted_files or (period_files if required_headers else [])
        candidates = []
        for path in scan_pool:
            try:
                sheet_name = select_sheet(path, required_headers)
            except RuntimeError:
                continue
            candidates.append((path, sheet_name))
        if len(candidates) != 1:
            if rule["level"] == "core":
                raise RuntimeError(f"{rule['label']}无法唯一匹配，候选={[path.name for path, _ in candidates]}")
            logger.warning("%s无法唯一匹配，按缺失处理", rule["label"])
            sources[role] = {"role": role, "label": rule["label"], "required_level": rule["level"], "status": "missing"}
            continue
        path, sheet_name = candidates[0]
        sources[role] = {
            "role": role,
            "label": rule["label"],
            "required_level": rule["level"],
            "status": "ready",
            "path": path,
            "file_name": path.name,
            "sheet_name": sheet_name,
        }
        logger.info("已定位%s：%s / %s", rule["label"], path.name, sheet_name)
    return sources


def read_rows(source: dict[str, Any]) -> list[dict[str, Any]]:
    """读取已确认工作表的所有数据行。"""

    if source.get("status") != "ready":
        return []
    workbook = load_workbook(source["path"], read_only=False, data_only=True)
    sheet = workbook[source["sheet_name"]]
    headers = [clean_text(sheet.cell(1, column).value) for column in range(1, sheet.max_column + 1)]
    rows = []
    for row_index in range(2, sheet.max_row + 1):
        values = [sheet.cell(row_index, column).value for column in range(1, sheet.max_column + 1)]
        row = dict(zip(headers, values))
        row["_row_index"] = row_index
        rows.append(row)
    workbook.close()
    logger.info("已读取%s：%s 行", source["label"], len(rows))
    return rows


def format_number(value: float | None, digits: int = 2) -> str:
    """格式化数量或金额差距。"""

    if value is None:
        return "-"
    return f"{value:.{digits}f}"


def format_interval_text(raw: str) -> str:
    """把区间文案中的数字统一为两位小数，并移除千位分隔符。"""

    def replace_number(match: re.Match[str]) -> str:
        number = float(match.group("number").replace(",", ""))
        return f"{format_number(number)}{match.group('scale')}"

    return re.sub(
        r"(?P<number>\d[\d,]*(?:\.\d+)?)(?P<scale>万?)",
        replace_number,
        raw,
    )


def ratio_label(self_value: float | None, competitor_value: float | None) -> str:
    """生成领先方倍率文案。"""

    if self_value is None or competitor_value is None or min(self_value, competitor_value) <= 0:
        return "无可比倍率"
    if self_value >= competitor_value:
        return f"本品 {self_value / competitor_value:.2f}x"
    return f"竞品 {competitor_value / self_value:.2f}x"


def gap_text(label: str, self_value: float | None, competitor_value: float | None) -> str:
    """生成网页统一使用的领先或落后文案。"""

    if self_value is None or competitor_value is None:
        return "数据不足"
    direction = "领先" if self_value >= competitor_value else "落后"
    if label == "成交转化率":
        gap = f"{abs(self_value - competitor_value) * 100:.2f}pct"
    else:
        gap = format_number(abs(self_value - competitor_value))
    return f"本品{direction} {gap} | {ratio_label(self_value, competitor_value)}"


def analyze_core(
    self_row: dict[str, Any],
    core_row: dict[str, Any],
    traffic_rows: list[dict[str, Any]],
    competitor_prefix: str,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], float | None, list[str]]:
    """生成核心指标估算与审计结果。

    功能说明：计算本品 P、竞品候选值，并按成交公式和顶层成交客户区间校正。
    参数 self_row：本品真实 SPU 数据行。
    参数 core_row：竞品核心区间数据行。
    参数 traffic_rows：流量来源原始行。
    参数 competitor_prefix：目标竞品字段前缀。
    返回值：本品校验、竞品转换、完整对比、指标卡、竞品成交人数和风险列表。
    """

    ranges: dict[str, ParsedRange] = {}
    self_values: dict[str, float | None] = {}
    candidates: dict[str, float | None] = {}
    validation = []
    risks: list[str] = []
    for metric in CORE_METRICS:
        actual = to_number(self_row.get(metric.label))
        self_interval = parse_range(core_row.get(f"本品{metric.label}"))
        competitor_interval = parse_range(core_row.get(f"{competitor_prefix}{metric.label}"))
        valid = in_range(actual, self_interval)
        position = None
        if actual is not None and self_interval.low is not None and self_interval.high is not None:
            position = 0.5 if self_interval.high == self_interval.low else (actual - self_interval.low) / (self_interval.high - self_interval.low)
            position = max(0.0, min(1.0, position))
        candidate = None
        if position is not None and competitor_interval.low is not None and competitor_interval.high is not None:
            candidate = competitor_interval.low + (competitor_interval.high - competitor_interval.low) * position
        if candidate is None:
            candidate = competitor_interval.mid
        median_error = abs((self_interval.mid or 0) - actual) / actual if actual and self_interval.mid is not None else None
        validation.append(
            {
                "metric_id": metric.id,
                "metric_label": metric.label,
                "actual_value": actual,
                "range_text": format_interval_text(self_interval.raw),
                "range_low": self_interval.low,
                "range_high": self_interval.high,
                "position_p": position,
                "in_range": valid,
                "median_error_rate": median_error,
                "historical_p": None,
                "historical_p_period": None,
                "historical_p_error_rate": None,
                "note": "本品使用真实 SPU 数据。" if valid else "本品真实值未落入本品区间。",
            }
        )
        if valid is False:
            risks.append(f"本品{metric.label}未落入本品区间")
        ranges[metric.id] = competitor_interval
        self_values[metric.id] = actual
        candidates[metric.id] = candidate

    final = dict(candidates)
    adjustments: list[str] = []
    top_customer_interval = ParsedRange(None, None, None, None, False, True)
    for row in traffic_rows:
        if clean_text(row.get("一级渠道")) and not clean_text(row.get("二级渠道")) and not clean_text(row.get("三级渠道")):
            top_customer_interval = parse_range(row.get(f"{competitor_prefix}成交客户数"))
            break

    visitors = final.get("visitors")
    conversion = final.get("conversion_rate")
    buyers = visitors * conversion if visitors is not None and conversion is not None else None
    if buyers is not None and not top_customer_interval.missing:
        target_buyers = clamp(buyers, top_customer_interval)
        if target_buyers != buyers and visitors:
            adjusted_rate = target_buyers / visitors
            if in_range(adjusted_rate, ranges["conversion_rate"]):
                final["conversion_rate"] = adjusted_rate
                conversion = adjusted_rate
                buyers = target_buyers
                adjustments.append("成交转化率按顶层成交客户数区间校正")
            elif conversion:
                adjusted_visitors = target_buyers / conversion
                if in_range(adjusted_visitors, ranges["visitors"]):
                    final["visitors"] = adjusted_visitors
                    visitors = adjusted_visitors
                    buyers = target_buyers
                    adjustments.append("访客数按顶层成交客户数区间校正")

    buyers = visitors * conversion if visitors is not None and conversion is not None else buyers
    price = final.get("customer_price")
    formula_gmv = buyers * price if buyers is not None and price is not None else final.get("gmv")
    if formula_gmv is not None:
        target_gmv = clamp(formula_gmv, ranges["gmv"])
        if buyers:
            adjusted_price = target_gmv / buyers
            if in_range(adjusted_price, ranges["customer_price"]):
                if price != adjusted_price:
                    adjustments.append("成交客单价按成交金额公式校正")
                final["customer_price"] = adjusted_price
                price = adjusted_price
        final["gmv"] = target_gmv
        if target_gmv != formula_gmv:
            adjustments.append("成交金额截断到原始区间")

    orders = final.get("orders")
    sold_units = final.get("sold_units")
    if orders is not None and sold_units is not None and sold_units < orders:
        final["sold_units"] = clamp(orders, ranges["sold_units"])
        adjustments.append("成交商品件数按不小于成交单量校正")

    conversions = []
    comparison = []
    cards = []
    for metric in CORE_METRICS:
        interval = ranges[metric.id]
        value = final.get(metric.id)
        self_value = self_values[metric.id]
        adjusted = value != candidates[metric.id]
        conversions.append(
            {
                "metric_id": metric.id,
                "metric_label": metric.label,
                "range_text": format_interval_text(interval.raw),
                "range_low": interval.low,
                "range_high": interval.high,
                "median_candidate": interval.mid,
                "p_candidate": candidates[metric.id],
                "historical_p_candidate": None,
                "final_value": value,
                "basis": "成交公式与流量约束校正" if adjusted else "同周期本品 P 候选",
                "confidence": "high" if in_range(value, interval) else "medium",
                "checks": {
                    "range_consistent": in_range(value, interval),
                    "traffic_consistent": in_range(buyers, top_customer_interval) if not top_customer_interval.missing else None,
                    "adjustments": adjustments,
                },
            }
        )
        status = "advantage" if self_value is not None and value is not None and self_value >= value else "warning"
        comparison.append(
            {
                "metric_id": metric.id,
                "metric_label": metric.label,
                "self_value": self_value,
                "competitor_value": value,
                "gap": self_value - value if self_value is not None and value is not None else None,
                "gap_pct_point": (self_value - value) * 100 if metric.unit == "%" and self_value is not None and value is not None else None,
                "ratio": self_value / value if self_value and value else None,
                "judgement": "本品领先" if status == "advantage" else "本品落后",
            }
        )
        if metric.id in CORE_CARD_IDS:
            card_self = self_value * 100 if metric.unit == "%" and self_value is not None else self_value
            card_competitor = value * 100 if metric.unit == "%" and value is not None else value
            cards.append(
                {
                    "id": metric.id,
                    "label": metric.label,
                    "unit": metric.unit,
                    "self_value": card_self,
                    "competitor_value": card_competitor,
                    "gap_abs_text": format_number(abs((self_value or 0) - (value or 0)) * (100 if metric.unit == "%" else 1)),
                    "ratio_text": ratio_label(self_value, value),
                    "gap_text": gap_text(metric.label, self_value, value),
                    "status": status,
                    "priority": "高" if status == "warning" else "低",
                }
            )
    return validation, conversions, comparison, cards, buyers, risks


def enrich_traffic_visitor_rates(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """补充渠道访客的同层占比和总占比。

    功能说明：按同一父渠道分组计算兄弟节点访客占比，并整理源表披露的全渠道访客占比；源表总占比缺失时按一级渠道访客合计回算。
    参数 rows：已经完成层级路径、两侧访客值和源表访客占比解析的渠道行。
    返回值：补充本品与竞品同层访客占比、总访客占比后的渠道行列表。
    """

    sibling_totals: dict[tuple[str, ...], dict[str, float]] = {}
    root_totals = {"self_visitors": 0.0, "competitor_visitors": 0.0}
    for row in rows:
        levels = tuple(
            value
            for value in (row.get("level_1"), row.get("level_2"), row.get("level_3"))
            if value and value != "-"
        )
        parent_key = levels[:-1]
        totals = sibling_totals.setdefault(parent_key, {"self_visitors": 0.0, "competitor_visitors": 0.0})
        for visitor_key in ("self_visitors", "competitor_visitors"):
            value = row.get(visitor_key)
            if isinstance(value, (int, float)):
                totals[visitor_key] += value
                if len(levels) == 1:
                    root_totals[visitor_key] += value

    for row in rows:
        levels = tuple(
            value
            for value in (row.get("level_1"), row.get("level_2"), row.get("level_3"))
            if value and value != "-"
        )
        totals = sibling_totals.get(levels[:-1], {})
        for side in ("self", "competitor"):
            visitor_key = f"{side}_visitors"
            source_rate_key = f"{side}_visitor_rate"
            value = row.get(visitor_key)
            sibling_total = totals.get(visitor_key)
            total_value = root_totals.get(visitor_key)
            current_level_rate = value / sibling_total if isinstance(value, (int, float)) and sibling_total else None
            source_total_rate = row.get(source_rate_key)
            total_rate = (
                source_total_rate
                if isinstance(source_total_rate, (int, float))
                else value / total_value if isinstance(value, (int, float)) and total_value else None
            )
            row[f"{side}_current_level_visitor_rate"] = current_level_rate
            row[f"{side}_total_visitor_rate"] = total_rate
    return rows


def analyze_traffic(rows: list[dict[str, Any]], competitor_prefix: str) -> list[dict[str, Any]]:
    """解析完整流量来源并生成渠道差距。

    功能说明：读取三级渠道区间数据，生成路径、两侧访客与成交估算、差距判断、转化效率及访客占比。
    参数 rows：流量来源工作表的标准化原始行。
    参数 competitor_prefix：源表中目标竞品字段使用的列名前缀。
    返回值：包含层级结构、对比指标、占比和建议动作的渠道明细列表。
    """

    result = []
    for row in rows:
        levels = [clean_text(row.get(key)) for key in ("一级渠道", "二级渠道", "三级渠道")]
        path_parts = [value for value in levels if value and value != "-"]
        if not path_parts:
            continue
        self_visitors = parse_range(row.get("本品访客数")).mid
        competitor_visitors = parse_range(row.get(f"{competitor_prefix}访客数")).mid
        self_gmv = parse_range(row.get("本品成交金额")).mid
        competitor_gmv = parse_range(row.get(f"{competitor_prefix}成交金额")).mid
        self_rate = parse_range(row.get("本品成交转化率")).mid
        competitor_rate = parse_range(row.get(f"{competitor_prefix}成交转化率")).mid
        visitor_gap = (self_visitors or 0) - (competitor_visitors or 0)
        gmv_gap = (self_gmv or 0) - (competitor_gmv or 0)
        rate_gap = ((self_rate or 0) - (competitor_rate or 0)) * 100
        if visitor_gap >= 0 and gmv_gap >= 0:
            judgement = "本品领先"
        elif visitor_gap < 0 and gmv_gap < 0:
            judgement = "本品落后"
        else:
            judgement = "结构分化"
        if visitor_gap >= 0 and rate_gap < 0:
            action = "优化价格权益、主图卖点、评价和促销承接"
        elif visitor_gap < 0:
            action = "补强渠道入口、资源位和投放承接"
        else:
            action = "保持渠道优势，沉淀可复用打法"
        result.append(
            {
                "level_1": levels[0] or None,
                "level_2": levels[1] or None,
                "level_3": levels[2] or None,
                "path": " > ".join(path_parts),
                "self_visitors": self_visitors,
                "competitor_visitors": competitor_visitors,
                "self_visitor_rate": parse_range(row.get("本品访客数占比")).mid,
                "competitor_visitor_rate": parse_range(row.get(f"{competitor_prefix}访客数占比")).mid,
                "self_gmv": self_gmv,
                "competitor_gmv": competitor_gmv,
                "self_conversion_rate": self_rate,
                "competitor_conversion_rate": competitor_rate,
                "self_customers": parse_range(row.get("本品成交客户数")).mid,
                "competitor_customers": parse_range(row.get(f"{competitor_prefix}成交客户数")).mid,
                "visitor_gap": visitor_gap,
                "gmv_gap": gmv_gap,
                "conversion_gap_pct": rate_gap,
                "judgement": judgement,
                "suggested_action": action,
                "estimation_basis": "区间中位数；核心指标另受顶层渠道约束",
            }
        )
    return enrich_traffic_visitor_rates(result)


def analyze_keywords(
    rows: list[dict[str, Any]],
    self_spu: str,
    competitor_spu: str,
    self_visitors: float | None,
    self_gmv: float | None,
    competitor_visitors: float | None,
    competitor_gmv: float | None,
) -> dict[str, Any]:
    """生成完整关键词外连接、覆盖率和机会判断。"""

    grouped: dict[str, dict[str, dict[str, Any]]] = {"self": {}, "competitor": {}}
    for row in rows:
        spu = clean_identifier(row.get("SPUID"))
        side = "self" if spu == self_spu else "competitor" if spu == competitor_spu else None
        keyword = clean_text(row.get("关键词"))
        if side and keyword:
            grouped[side][keyword.casefold()] = row
    keys = sorted(set(grouped["self"]) | set(grouped["competitor"]))
    result_rows = []
    for key in keys:
        self_row = grouped["self"].get(key)
        competitor_row = grouped["competitor"].get(key)
        display_keyword = clean_text((self_row or competitor_row or {}).get("关键词"))
        self_keyword_visitors = parse_range(self_row.get("访客数") if self_row else None).mid
        competitor_keyword_visitors = parse_range(competitor_row.get("访客数") if competitor_row else None).mid
        self_keyword_gmv = parse_range(self_row.get("成交金额") if self_row else None).mid
        competitor_keyword_gmv = parse_range(competitor_row.get("成交金额") if competitor_row else None).mid
        relation = "共同词" if self_row and competitor_row else "本品独有" if self_row else "竞品独有"
        visitor_gap = (self_keyword_visitors or 0) - (competitor_keyword_visitors or 0)
        gmv_gap = (self_keyword_gmv or 0) - (competitor_keyword_gmv or 0)
        if relation == "竞品独有":
            opportunity = "补词机会"
        elif relation == "本品独有":
            opportunity = "保持优势"
        elif visitor_gap < 0:
            opportunity = "访客落后"
        elif gmv_gap < 0:
            opportunity = "成交落后"
        else:
            opportunity = "本品领先"
        result_rows.append(
            {
                "keyword": display_keyword,
                "coverage_relation": relation,
                "self_visitors": self_keyword_visitors,
                "competitor_visitors": competitor_keyword_visitors,
                "visitor_gap": visitor_gap,
                "self_gmv": self_keyword_gmv,
                "competitor_gmv": competitor_keyword_gmv,
                "gmv_gap": gmv_gap,
                "opportunity": opportunity,
            }
        )
    result_rows.sort(key=lambda item: max(abs(item["gmv_gap"]), abs(item["visitor_gap"])), reverse=True)
    self_visitor_sum = sum(item["self_visitors"] or 0 for item in result_rows)
    competitor_visitor_sum = sum(item["competitor_visitors"] or 0 for item in result_rows)
    self_gmv_sum = sum(item["self_gmv"] or 0 for item in result_rows)
    competitor_gmv_sum = sum(item["competitor_gmv"] or 0 for item in result_rows)
    return {
        "summary": {
            "common_count": sum(item["coverage_relation"] == "共同词" for item in result_rows),
            "self_only_count": sum(item["coverage_relation"] == "本品独有" for item in result_rows),
            "competitor_only_count": sum(item["coverage_relation"] == "竞品独有" for item in result_rows),
        },
        "coverage": {
            "self_visitor_rate": self_visitor_sum / self_visitors if self_visitors else None,
            "competitor_visitor_rate": competitor_visitor_sum / competitor_visitors if competitor_visitors else None,
            "self_gmv_rate": self_gmv_sum / self_gmv if self_gmv else None,
            "competitor_gmv_rate": competitor_gmv_sum / competitor_gmv if competitor_gmv else None,
        },
        "rows": result_rows,
        "notes": ["关键词列表为 Top 口径，合计不代表商品全量。"],
    }


def analyze_profile(rows: list[dict[str, Any]], competitor_prefix: str, self_buyers: float | None, competitor_buyers: float | None) -> dict[str, Any]:
    """解析成交客户画像的维度标题与完整画像项。"""

    dimensions: list[dict[str, Any]] = []
    current: dict[str, Any] | None = None
    for row in rows:
        name = clean_text(row.get("画像类型"))
        self_interval = parse_range(row.get("本品成交客户数占比"))
        competitor_interval = parse_range(row.get(f"{competitor_prefix}成交客户数占比"))
        if name in DIMENSION_NAMES and self_interval.missing and competitor_interval.missing:
            current = {"dimension": name, "items": []}
            dimensions.append(current)
            continue
        if not name or current is None:
            continue
        self_rate = self_interval.mid * 100 if self_interval.mid is not None else None
        competitor_rate = competitor_interval.mid * 100 if competitor_interval.mid is not None else None
        gap_rate = self_rate - competitor_rate if self_rate is not None and competitor_rate is not None else None
        if gap_rate is None:
            judgement = "无完整口径"
        elif gap_rate >= 1:
            judgement = "本品领先"
        elif gap_rate <= -1:
            judgement = "本品落后"
        else:
            judgement = "基本持平"
        current["items"].append(
            {
                "dimension": current["dimension"],
                "name": name,
                "self_rate": self_rate,
                "competitor_rate": competitor_rate,
                "gap_rate": gap_rate,
                "self_estimated_customers": self_buyers * self_rate / 100 if self_buyers is not None and self_rate is not None else None,
                "competitor_estimated_customers": competitor_buyers * competitor_rate / 100 if competitor_buyers is not None and competitor_rate is not None else None,
                "judgement": judgement,
            }
        )
    return {"dimensions": dimensions, "notes": ["各画像维度分别比较；缺失值不等同于 0。"]}


def analyze_promotion(rows: list[dict[str, Any]], competitor_prefix: str, competitor_gmv: float | None) -> dict[str, Any]:
    """解析推广点击、归因成交和贡献比例。"""

    if not rows:
        return {"available": False, "self": {}, "competitor": {}, "attributed_gmv_rate": None, "judgement": "数据不足", "notes": []}
    row = rows[0]
    self_non_site_clicks = parse_range(row.get("非全站-本店商品广告点击数")).mid
    self_non_site_gmv = parse_range(row.get("非全站-本店商品广告总订单金额")).mid
    competitor_non_site_clicks = parse_range(row.get(f"非全站-{competitor_prefix}广告点击数")).mid
    competitor_non_site_gmv = parse_range(row.get(f"非全站-{competitor_prefix}广告总订单金额")).mid
    attributed_rate = competitor_non_site_gmv / competitor_gmv if competitor_non_site_gmv is not None and competitor_gmv else None
    return {
        "available": True,
        "self": {"ad_clicks": self_non_site_clicks, "ad_order_gmv": self_non_site_gmv},
        "competitor": {"ad_clicks": competitor_non_site_clicks, "ad_order_gmv": competitor_non_site_gmv},
        "attributed_gmv_rate": attributed_rate,
        "judgement": "竞品投放贡献较高" if attributed_rate is not None and attributed_rate >= 0.3 else "竞品投放贡献有限",
        "notes": ["推广成交为广告归因口径，不与核心 SPU 成交金额强制对齐。"],
    }


def build_tabs(traffic: list[dict[str, Any]], keywords: dict[str, Any], profile: dict[str, Any]) -> list[dict[str, Any]]:
    """生成 HTML 直接消费的三个差距来源 Tab。"""

    def select_balanced_highlights(items: list[dict[str, Any]]) -> list[dict[str, Any]]:
        """优先各选一项优势和劣势，缺少任一类型时按原顺序补足两项。"""

        selected: list[dict[str, Any]] = []
        for status in ("advantage", "warning"):
            match = next((item for item in items if item["status"] == status), None)
            if match is not None:
                selected.append(match)
        for item in items:
            if len(selected) >= 2:
                break
            if item not in selected:
                selected.append(item)
        return selected

    def traffic_highlight_gap(item: dict[str, Any]) -> str:
        """生成人员可直接理解的渠道访客差距摘要。"""

        if item["self_visitors"] is None and item["competitor_visitors"] is not None:
            return f"竞品独有 | 竞品访客 {format_number(item['competitor_visitors'])}"
        if item["competitor_visitors"] is None and item["self_visitors"] is not None:
            return f"本品独有 | 本品访客 {format_number(item['self_visitors'])}"
        return gap_text("访客数", item["self_visitors"], item["competitor_visitors"])

    traffic_sorted = sorted(traffic, key=lambda item: abs(item["visitor_gap"]), reverse=True)
    traffic_highlight_candidates = [
        {
            "label": item["path"],
            "self_value": item["self_visitors"],
            "competitor_value": item["competitor_visitors"],
            "unit": "",
            "gap_text": traffic_highlight_gap(item),
            "status": "advantage" if item["visitor_gap"] >= 0 else "warning",
            "action": item["suggested_action"],
        }
        for item in traffic_sorted
    ]
    traffic_highlights = select_balanced_highlights(traffic_highlight_candidates)
    traffic_rows = [
        {
            **item,
            "self_conversion_rate_pct": item["self_conversion_rate"] * 100 if item["self_conversion_rate"] is not None else None,
            "competitor_conversion_rate_pct": item["competitor_conversion_rate"] * 100 if item["competitor_conversion_rate"] is not None else None,
            "self_current_level_visitor_rate_pct": item["self_current_level_visitor_rate"] * 100 if item["self_current_level_visitor_rate"] is not None else None,
            "competitor_current_level_visitor_rate_pct": item["competitor_current_level_visitor_rate"] * 100 if item["competitor_current_level_visitor_rate"] is not None else None,
            "self_total_visitor_rate_pct": item["self_total_visitor_rate"] * 100 if item["self_total_visitor_rate"] is not None else None,
            "competitor_total_visitor_rate_pct": item["competitor_total_visitor_rate"] * 100 if item["competitor_total_visitor_rate"] is not None else None,
        }
        for item in traffic
    ]

    keyword_rows = keywords["rows"]
    keyword_highlight_candidates = [
        {
            "label": item["keyword"],
            "self_value": item["self_visitors"],
            "competitor_value": item["competitor_visitors"],
            "unit": "",
            "gap_text": f"{item['opportunity']} | 访客差距 {format_number(item['visitor_gap'])} | 成交差距 {format_number(item['gmv_gap'])}",
            "status": "warning" if item["opportunity"] in {"补词机会", "访客落后", "成交落后"} else "advantage",
            "action": "补充搜索词并检查页面承接" if item["opportunity"] != "保持优势" else "保持优势词覆盖",
        }
        for item in keyword_rows
    ]
    keyword_highlights = select_balanced_highlights(keyword_highlight_candidates)

    profile_rows = [item for dimension in profile["dimensions"] for item in dimension["items"]]
    comparable_profile = [item for item in profile_rows if item["gap_rate"] is not None]
    comparable_profile.sort(key=lambda item: abs(item["gap_rate"]), reverse=True)
    profile_highlight_candidates = [
        {
            "label": f"{item['dimension']}：{item['name']}",
            "self_value": item["self_rate"],
            "competitor_value": item["competitor_rate"],
            "unit": "%",
            "gap_text": f"{item['judgement']} | 差距 {abs(item['gap_rate']):.2f}pct",
            "status": "warning" if item["judgement"] == "本品落后" else "advantage",
            "action": "结合该画像调整内容表达与素材占比",
        }
        for item in comparable_profile
    ]
    profile_highlights = select_balanced_highlights(profile_highlight_candidates)

    return [
        {
            "id": "traffic",
            "label": "流量来源",
            "headline": "按渠道路径比较访客规模、成交金额和转化效率",
            "highlights": traffic_highlights,
            "columns": [
                {"key": "path", "label": "渠道路径"},
                {"key": "judgement", "label": "判断"},
                {"key": "visitor_gap", "label": "访客差距"},
                {"key": "gmv_gap", "label": "成交金额差距"},
                {"key": "conversion_gap_pct", "label": "转化差距", "unit": "pct"},
                {"key": "self_current_level_visitor_rate_pct", "label": "本品同层访客占比", "unit": "%"},
                {"key": "competitor_current_level_visitor_rate_pct", "label": "竞品同层访客占比", "unit": "%"},
                {"key": "self_total_visitor_rate_pct", "label": "本品总访客占比", "unit": "%"},
                {"key": "competitor_total_visitor_rate_pct", "label": "竞品总访客占比", "unit": "%"},
                {"key": "self_visitors", "label": "本品访客"},
                {"key": "competitor_visitors", "label": "竞品访客"},
                {"key": "self_gmv", "label": "本品成交金额"},
                {"key": "competitor_gmv", "label": "竞品成交金额"},
                {"key": "self_conversion_rate_pct", "label": "本品转化率", "unit": "%"},
                {"key": "competitor_conversion_rate_pct", "label": "竞品转化率", "unit": "%"},
            ],
            "rows": traffic_rows,
            "notes": ["同层访客占比按同一父渠道下的兄弟节点计算；总访客占比优先采用源表披露值。渠道值按原始区间估算，核心指标另受顶层渠道约束。"],
        },
        {
            "id": "keywords",
            "label": "关键词",
            "headline": "比较全部披露关键词的覆盖关系、访客和成交差距",
            "highlights": keyword_highlights,
            "columns": [
                {"key": "keyword", "label": "关键词"},
                {"key": "opportunity", "label": "机会判断"},
                {"key": "visitor_gap", "label": "访客差距"},
                {"key": "gmv_gap", "label": "成交差距"},
                {"key": "coverage_relation", "label": "覆盖关系"},
                {"key": "self_visitors", "label": "本品访客"},
                {"key": "competitor_visitors", "label": "竞品访客"},
                {"key": "self_gmv", "label": "本品成交金额"},
                {"key": "competitor_gmv", "label": "竞品成交金额"},
            ],
            "rows": keyword_rows,
            "notes": keywords["notes"],
        },
        {
            "id": "customer_profile",
            "label": "客户画像",
            "headline": "按性别、年龄、地区、省份和城市维度比较成交客户结构",
            "highlights": profile_highlights,
            "dimension_field": "dimension",
            "dimension_label": "画像维度",
            "columns": [
                {"key": "name", "label": "画像项"},
                {"key": "judgement", "label": "判断"},
                {"key": "gap_rate", "label": "占比差距", "unit": "pct"},
                {"key": "self_rate", "label": "本品占比", "unit": "%"},
                {"key": "competitor_rate", "label": "竞品占比", "unit": "%"},
            ],
            "rows": profile_rows,
            "notes": profile["notes"],
        },
    ]


def build_diagnosis(cards: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """把核心指标差距转换为证据、建议和行动跟踪。"""

    recommendations = {
        "gmv": "拆解领先或落后的流量、转化和客单贡献，按周复盘成交金额。",
        "visitors": "保持优势渠道，补强落后渠道入口并检查资源位承接。",
        "conversion_rate": "核查价格权益、详情页首屏、评价晒图和搜索词承接。",
        "customer_price": "优化套装、满减和高客单 SKU 露出。",
    }
    diagnosis = []
    actions = []
    for card in cards:
        title_prefix = "优势" if card["status"] == "advantage" else "短板"
        recommendation = recommendations[card["id"]]
        diagnosis.append(
            {
                "title": f"{title_prefix}：{card['label']}",
                "text": card["gap_text"],
                "evidence": card["gap_text"],
                "recommendation": recommendation,
                "status": card["status"],
                "source": f"core_metrics.{card['id']}",
            }
        )
        actions.append(
            {
                "module": "核心指标",
                "opportunity": f"{card['label']}{'保持' if card['status'] == 'advantage' else '优化'}",
                "gap_basis": card["gap_text"],
                "suggested_action": recommendation,
                "priority": card["priority"],
                "review_metric": card["label"],
                "status": "待推进",
            }
        )
    return diagnosis, actions


def empty_contract() -> dict[str, Any]:
    """生成 HTML 可直接读取的空结构模板。"""

    empty_keywords = {
        "summary": {"common_count": 0, "self_only_count": 0, "competitor_only_count": 0},
        "coverage": {"self_visitor_rate": None, "competitor_visitor_rate": None, "self_gmv_rate": None, "competitor_gmv_rate": None},
        "rows": [],
        "notes": [],
    }
    empty_profile = {"dimensions": [], "notes": []}
    return {
        "schema_version": "1.0",
        "meta": {
            "title": None,
            "period": None,
            "period_start": None,
            "period_end": None,
            "period_key": None,
            "granularity": None,
            "self_name": None,
            "self_spu": None,
            "competitor_name": None,
            "competitor_spu": None,
            "confidence": None,
            "summary": None,
            "weakness_summary": None,
        },
        "source_files": [],
        "self_validation": [],
        "competitor_core_conversions": [],
        "core_metrics": [],
        "comparison": [],
        "traffic_sources": [],
        "keywords": empty_keywords,
        "customer_profile": empty_profile,
        "promotion": {"available": False, "self": {}, "competitor": {}, "attributed_gmv_rate": None, "judgement": None, "notes": []},
        "tabs": build_tabs([], empty_keywords, empty_profile),
        "ai_recommendations": [],
        "diagnosis": [],
        "action_tracking": [],
        "risks": [],
    }


def build_analysis_result(args: argparse.Namespace) -> tuple[dict[str, Any], dict[str, Any]]:
    """从真实原始 Excel 生成当前 HTML 契约。

    功能说明：发现六类源文件，完成核心估算、差距分析和页面字段组装。
    参数 args：命令行参数，包含输入目录、周期、粒度、SPU 和竞品前缀。
    返回值：完整 `analysis_result.json` 与 `normalized_data.json` 字典。
    """

    sources = discover_sources(Path(args.input_dir), args.period_file, args.competitor_prefix)
    data_rows = {role: read_rows(source) for role, source in sources.items()}
    self_rows = data_rows["self_real"]
    self_matches = [row for row in self_rows if clean_identifier(row.get("SPU")) == args.self_spu]
    if len(self_matches) != 1:
        raise RuntimeError(f"本品 SPU 无法唯一匹配：{args.self_spu}，匹配行数={len(self_matches)}")
    self_row = self_matches[0]
    if not data_rows["core"]:
        raise RuntimeError("竞品核心数据对比没有数据行")
    core_row = data_rows["core"][0]

    validation, conversions, comparison, cards, competitor_buyers, risks = analyze_core(
        self_row, core_row, data_rows["traffic"], args.competitor_prefix
    )
    conversion_map = {item["metric_id"]: item["final_value"] for item in conversions}
    self_value_map = {item["metric_id"]: item["actual_value"] for item in validation}
    traffic = analyze_traffic(data_rows["traffic"], args.competitor_prefix)
    keywords = analyze_keywords(
        data_rows["keywords"],
        args.self_spu,
        args.competitor_spu,
        self_value_map.get("visitors"),
        self_value_map.get("gmv"),
        conversion_map.get("visitors"),
        conversion_map.get("gmv"),
    )
    profile = analyze_profile(data_rows["customer_profile"], args.competitor_prefix, to_number(self_row.get("成交人数")), competitor_buyers)
    promotion = analyze_promotion(data_rows["promotion"], args.competitor_prefix, conversion_map.get("gmv"))
    tabs = build_tabs(traffic, keywords, profile)
    diagnosis, actions = build_diagnosis(cards)

    keyword_rows = data_rows["keywords"]
    competitor_names = {
        clean_text(row.get("商品名称"))
        for row in keyword_rows
        if clean_identifier(row.get("SPUID")) == args.competitor_spu and clean_text(row.get("商品名称"))
    }
    advantage_labels = [item["label"] for item in cards if item["status"] == "advantage"]
    warning_labels = [item["label"] for item in cards if item["status"] == "warning"]
    if sources["keywords"].get("status") != "ready":
        risks.append("缺少关键词数据，关键词 Tab 不完整")
    if sources["customer_profile"].get("status") != "ready":
        risks.append("缺少客户画像数据，客户画像 Tab 不完整")
    if not promotion["available"]:
        risks.append("缺少推广数据，投放判断不可用")
    risks.append("竞品数值为区间数据按当前 SOP 生成的准真实估算值")

    source_files = [
        {
            "role": source["role"],
            "label": source["label"],
            "file_name": source.get("file_name"),
            "sheet_name": source.get("sheet_name"),
            "required_level": source["required_level"],
            "status": source["status"],
            "warnings": [],
        }
        for source in sources.values()
    ]
    generated_at = datetime.now().isoformat(timespec="seconds")
    period_meta = period_fields(args.granularity, args.period)
    normalized = {
        "schema_version": "1.0",
        "meta": {
            **period_meta,
            "period_file": args.period_file,
            "granularity": args.granularity,
            "self_spu": args.self_spu,
            "competitor_spu": args.competitor_spu,
            "competitor_prefix": args.competitor_prefix,
            "generated_at": generated_at,
        },
        "source_files": source_files,
        "core_raw": core_row,
        "self_real": self_row,
        "traffic_rows": data_rows["traffic"],
        "keyword_rows": data_rows["keywords"],
        "customer_profile_rows": data_rows["customer_profile"],
        "promotion_rows": data_rows["promotion"],
        "warnings": [risk for risk in risks if "缺少" in risk],
    }
    result = {
        "schema_version": "1.0",
        "meta": {
            "title": args.title,
            **period_meta,
            "granularity": args.granularity,
            "self_name": clean_text(self_row.get("商品名称")) or None,
            "self_spu": args.self_spu,
            "competitor_name": sorted(competitor_names)[0] if competitor_names else None,
            "competitor_spu": args.competitor_spu,
            "confidence": "high" if all(item["in_range"] for item in validation) else "medium",
            "summary": f"本品{'、'.join(advantage_labels)}领先。" if advantage_labels else "本品核心指标暂无明显优势。",
            "weakness_summary": f"本品{'、'.join(warning_labels)}落后，需要优先优化。" if warning_labels else "本品核心指标暂无明显短板。",
            "generated_at": generated_at,
        },
        "source_files": source_files,
        "self_validation": validation,
        "competitor_core_conversions": conversions,
        "core_metrics": cards,
        "comparison": comparison,
        "traffic_sources": traffic,
        "keywords": keywords,
        "customer_profile": profile,
        "promotion": promotion,
        "tabs": tabs,
        "ai_recommendations": [],
        "diagnosis": diagnosis,
        "action_tracking": actions,
        "risks": list(dict.fromkeys(risks)),
    }
    return result, normalized


def validate_contract(data: dict[str, Any], allow_empty: bool = False) -> None:
    """校验 JSON 是否满足当前 HTML 和审计契约。

    功能说明：检查顶层模块、三个 Tab、指标卡和建议字段，失败时阻止写出。
    参数 data：待校验的分析结果字典。
    参数 allow_empty：是否允许指标和明细为空，用于空结构模板。
    返回值：无；契约不完整时抛出 `ValueError`。
    """

    required_top = {
        "schema_version",
        "meta",
        "source_files",
        "self_validation",
        "competitor_core_conversions",
        "core_metrics",
        "comparison",
        "traffic_sources",
        "keywords",
        "customer_profile",
        "promotion",
        "tabs",
        "ai_recommendations",
        "diagnosis",
        "action_tracking",
        "risks",
    }
    missing = sorted(required_top - set(data))
    if missing:
        raise ValueError(f"analysis_result 缺少顶层字段：{missing}")
    required_meta = {"period", "period_start", "period_end", "period_key", "granularity"}
    missing_meta = sorted(required_meta - set(data["meta"]))
    if missing_meta:
        raise ValueError(f"analysis_result.meta 缺少周期字段：{missing_meta}")
    if not isinstance(data["risks"], list) or any(not isinstance(item, str) for item in data["risks"]):
        raise ValueError("risks 必须是字符串数组")
    if not isinstance(data["ai_recommendations"], list):
        raise ValueError("ai_recommendations 必须是数组")
    for item in data["ai_recommendations"]:
        for field in ("source_id", "source_label", "target", "status", "evidence", "actions", "validation"):
            if field not in item:
                raise ValueError(f"AI 建议缺少字段：{field}")
        if not isinstance(item["actions"], list) or not item["actions"]:
            raise ValueError("AI 建议 actions 必须是非空数组")
    tab_map = {tab.get("id"): tab for tab in data["tabs"]}
    if set(tab_map) != {"traffic", "keywords", "customer_profile"}:
        raise ValueError(f"tabs 必须包含 traffic、keywords、customer_profile，当前={sorted(tab_map)}")
    for tab_id, tab in tab_map.items():
        for field in ("label", "headline", "highlights", "columns", "rows", "notes"):
            if field not in tab:
                raise ValueError(f"Tab {tab_id} 缺少字段：{field}")
    if not allow_empty:
        if len(data["core_metrics"]) != 4:
            raise ValueError("正式结果必须包含四张核心指标卡")
        for item in data["core_metrics"]:
            for field in ("id", "label", "unit", "self_value", "competitor_value", "gap_text", "status"):
                if field not in item:
                    raise ValueError(f"核心指标卡缺少字段：{field}")
        for item in data["diagnosis"]:
            for field in ("title", "evidence", "recommendation", "status", "source"):
                if field not in item:
                    raise ValueError(f"诊断建议缺少字段：{field}")
    logger.info("JSON 契约校验通过：allow_empty=%s", allow_empty)


def write_json(path: Path, data: dict[str, Any]) -> None:
    """以 UTF-8 和缩进格式原子写入 JSON。"""

    path.parent.mkdir(parents=True, exist_ok=True)
    temp_path = path.with_suffix(f"{path.suffix}.tmp")
    temp_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    temp_path.replace(path)
    logger.info("已写入 JSON：%s", path)


def write_html(path: Path, template_path: Path, data: dict[str, Any]) -> None:
    """把分析 JSON 内嵌到 HTML 模板并写入报告。"""

    template = template_path.read_text(encoding="utf-8")
    inline_data = json.dumps(data, ensure_ascii=False).replace("</", "<\\/")
    script = f"<script>window.__ANALYSIS_DATA__ = {inline_data};</script>\n"
    html = template.replace("</head>", script + "</head>", 1)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(html, encoding="utf-8")
    logger.info("已写入 HTML：%s", path)


def parse_args() -> argparse.Namespace:
    """解析真实 Excel 到网页看板的命令行参数。"""

    parser = argparse.ArgumentParser(description="读取京东真实原始 Excel，生成标准 analysis_result.json 和 HTML 看板。")
    parser.add_argument("--batch", action="store_true", help="扫描日、周、月目录并批量生成全部周期。")
    parser.add_argument("--input-root", help="批量模式的原始数据根目录，目录下包含天、周、月。")
    parser.add_argument("--input-dir", help="单周期模式的原始 Excel 目录。")
    parser.add_argument("--period-file", help="单周期模式的文件名周期片段，例如 YYYY-MM-DD_YYYY-MM-DD。")
    parser.add_argument("--period", help="单周期模式的页面展示周期，例如 YYYY-MM-DD~YYYY-MM-DD。")
    parser.add_argument("--granularity", choices=["day", "week", "month"], help="单周期模式的分析粒度。")
    parser.add_argument("--self-spu", required=True, help="本品 SPU。")
    parser.add_argument("--competitor-spu", required=True, help="竞品 SPU。")
    parser.add_argument("--competitor-prefix", default="竞品1", help="导出表中的目标竞品字段前缀。")
    parser.add_argument("--title", default="竞品准真实值看板", help="网页标题。")
    parser.add_argument("--output-root", help="批量模式的输出根目录。")
    parser.add_argument("--start-date", help="批量模式的最早周期日期，格式为 YYYY-MM-DD。")
    parser.add_argument("--end-date", help="批量模式的最晚周期日期，格式为 YYYY-MM-DD。")
    parser.add_argument("--output-json", help="单周期模式的 analysis_result.json 输出路径。")
    parser.add_argument("--output-normalized", help="可选的 normalized_data.json 输出路径。")
    parser.add_argument("--output-html", help="可选的内置数据 HTML 输出路径。")
    parser.add_argument("--empty-template-output", help="可选的空结构 JSON 输出路径。")
    parser.add_argument(
        "--template",
        default=str(Path(__file__).resolve().parents[1] / "assets" / "dashboard-template.html"),
        help="HTML 模板路径。",
    )
    parser.add_argument("--log-level", default="INFO", help="日志级别。")
    return parser.parse_args()


def run_single(args: argparse.Namespace) -> None:
    """执行单周期分析。

    功能说明：读取一个指定粒度和周期，生成分析结果、标准化数据及可选 HTML。
    参数 args：包含输入目录、周期、粒度、商品和输出路径的命令行参数。
    返回值：无；成功时写入调用方指定文件。
    """

    required = {
        "input_dir": args.input_dir,
        "period_file": args.period_file,
        "period": args.period,
        "granularity": args.granularity,
        "output_json": args.output_json,
    }
    missing = [name for name, value in required.items() if not value]
    if missing:
        raise ValueError(f"单周期模式缺少参数：{', '.join(missing)}")
    logger.info("开始生成看板数据：%s / %s", args.period, args.granularity)
    result, normalized = build_analysis_result(args)
    validate_contract(result)
    write_json(Path(args.output_json), result)
    if args.output_normalized:
        write_json(Path(args.output_normalized), normalized)
    if args.empty_template_output:
        empty = empty_contract()
        validate_contract(empty, allow_empty=True)
        write_json(Path(args.empty_template_output), empty)
    if args.output_html:
        write_html(Path(args.output_html), Path(args.template), result)
    logger.info("单周期看板数据生成完成")


def run_batch(args: argparse.Namespace) -> None:
    """执行日、周、月多周期批量分析。

    功能说明：分别扫描三个粒度目录，以核心对比文件发现周期，每周期独立生成两份 JSON，最后生成轻量报告索引。
    参数 args：包含输入根目录、输出根目录、商品信息和可选日期窗口的命令行参数。
    返回值：无；成功时写入周期结果和 `report-index.json`。
    """

    if not args.input_root or not args.output_root:
        raise ValueError("批量模式必须提供 --input-root 和 --output-root")
    for field_name, value in (("start_date", args.start_date), ("end_date", args.end_date)):
        if value:
            try:
                datetime.strptime(value, "%Y-%m-%d")
            except ValueError as error:
                raise ValueError(f"{field_name} 必须使用 YYYY-MM-DD 格式：{value}") from error
    if args.start_date and args.end_date and args.start_date > args.end_date:
        raise ValueError("start_date 不能晚于 end_date")
    input_root = Path(args.input_root).resolve()
    output_root = Path(args.output_root).resolve()
    generated_at = datetime.now().isoformat(timespec="seconds")
    report_index: dict[str, Any] = {
        "schema_version": "1.0",
        "updated_at": generated_at,
        "meta": {
            "title": args.title,
            "self_spu": args.self_spu,
            "competitor_spu": args.competitor_spu,
        },
        "reports": {granularity: [] for granularity in GRANULARITY_DIRS},
    }

    for granularity, directory_name in GRANULARITY_DIRS.items():
        input_dir = input_root / directory_name
        if not input_dir.is_dir():
            logger.warning("粒度目录不存在，跳过：%s", input_dir)
            continue
        for period_start, period_end in discover_periods(input_dir):
            if not period_in_window(period_start, period_end, args.start_date, args.end_date):
                continue
            period_file = f"{period_start}_{period_end}"
            period_meta = period_fields(granularity, period_file)
            period_args = argparse.Namespace(**vars(args))
            period_args.input_dir = str(input_dir)
            period_args.period_file = period_file
            period_args.period = period_meta["period"]
            period_args.granularity = granularity
            period_dir = output_root / granularity / period_file
            logger.info("开始处理周期：%s", period_meta["period_key"])
            result, normalized = build_analysis_result(period_args)
            validate_contract(result)
            write_json(period_dir / "normalized_data.json", normalized)
            write_json(period_dir / "analysis_result.json", result)
            report_index["reports"][granularity].append(
                {
                    **period_meta,
                    "generated_at": result["meta"]["generated_at"],
                    "confidence": result["meta"]["confidence"],
                    "path": f"/reports/{granularity}/{period_file}/analysis_result.json",
                }
            )
            logger.info("周期处理完成：%s", period_meta["period_key"])

    write_json(output_root / "report-index.json", report_index)
    counts = {key: len(value) for key, value in report_index["reports"].items()}
    logger.info("批量分析完成：%s", counts)


def main() -> None:
    """执行真实 Excel 到网页看板的数据管线。

    功能说明：根据命令行模式执行单周期分析或日、周、月批量分析。
    参数：无，参数由命令行提供。
    返回值：无；成功时写入调用方指定的输出文件。
    """

    args = parse_args()
    logging.basicConfig(level=getattr(logging, args.log_level.upper(), logging.INFO), format="%(levelname)s %(name)s - %(message)s")
    if args.batch:
        run_batch(args)
    else:
        run_single(args)


if __name__ == "__main__":
    main()
